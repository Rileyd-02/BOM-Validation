import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st
import io

st.title("📊 Excel BOM Comparison Tool")

# --- File uploader ---
uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file:
    # --- Load Excel sheets ---
    plm = pd.read_excel(uploaded_file, sheet_name="PLM")
    sap = pd.read_excel(uploaded_file, sheet_name="SAP")

    # --- Combine PLM ---
    plm["Combined"] = (
        plm["Material"].astype(str).str.strip()
        + "_" + plm["Vendor Reference"].astype(str).str.strip()
        + "_" + plm["Color Reference"].astype(str).str.strip()
    )

    # --- Combine SAP ---
    sap["Combined"] = (
        sap["Material"].astype(str).str.strip()
        + "_" + sap["Vendor Reference"].astype(str).str.strip()
        + "_" + sap["Comp. Colour"].astype(str).str.strip()
    )

    # --- Compare by join ---
    comparison = pd.merge(
        plm,
        sap,
        on="Combined",
        how="inner",
        suffixes=("_PLM", "_SAP")
    )

    # --- Extract relevant columns ---
    comparison_selected = comparison[
        [
            "Combined",
            "Material_PLM",
            "Material_SAP",
            "Vendor Reference_PLM",
            "Vendor Reference_SAP",
            "Color Reference",
            "Comp. Colour",
            "Qty(Cons.)",
            "Comp.Qty."
        ]
    ].copy()

    # --- Add consumption difference ---
    comparison_selected["Consumption Difference"] = (
        (comparison_selected["Qty(Cons.)"] - comparison_selected["Comp.Qty."]).abs()
    )

    # --- Sort by difference ---
    comparison_selected = comparison_selected.sort_values(
        by="Consumption Difference", ascending=False
    )

    # --- Also keep unmatched rows ---
    not_in_sap = plm[~plm["Combined"].isin(sap["Combined"])]
    not_in_plm = sap[~sap["Combined"].isin(plm["Combined"])]

    # --- Save results to BytesIO buffer instead of file ---
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        comparison_selected.to_excel(writer, sheet_name="Comparison", index=False)
        not_in_sap.to_excel(writer, sheet_name="PLM_Not_in_SAP", index=False)
        not_in_plm.to_excel(writer, sheet_name="SAP_Not_in_PLM", index=False)

    # --- Apply conditional formatting ---
    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Comparison"]

    headers = [cell.value for cell in ws[1]]
    plm_col = headers.index("Qty(Cons.)") + 1
    sap_col = headers.index("Comp.Qty.") + 1
    diff_col = headers.index("Consumption Difference") + 1

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        diff_val = ws.cell(row=row, column=diff_col).value
        if diff_val == 0:
            ws.cell(row=row, column=plm_col).fill = green_fill
            ws.cell(row=row, column=sap_col).fill = green_fill
            ws.cell(row=row, column=diff_col).fill = green_fill
        else:
            ws.cell(row=row, column=plm_col).fill = red_fill
            ws.cell(row=row, column=sap_col).fill = red_fill
            ws.cell(row=row, column=diff_col).fill = red_fill

    # Save again into buffer
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # --- Streamlit UI ---
    st.success("✅ Comparison complete!")
    st.subheader("Preview of Comparison Results")
    st.dataframe(comparison_selected.head(20))

    # --- Download button ---
    st.download_button(
        label="⬇️ Download Results (Excel)",
        data=final_output,
        file_name="comparison_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
